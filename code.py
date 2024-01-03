import openpyxl
from openpyxl import Workbook
import time

# Load the Excel workbook
workbook = openpyxl.load_workbook('data.xlsx')

# Assuming the data is in the first sheet, change it accordingly if it's in a different sheet
sheet = workbook.active
flnames = sheet.cell(row=3, column=4).value
brands = sheet.cell(row=3, column=3).value
print(type(flnames))
print(brands)
# time.sleep(11111)
formatted_emails = []

# Iterate through rows, starting from the second row (assuming the first row contains headers)
for row in sheet.iter_rows(min_row=2, values_only=True, max_col=1):
    # Extract first and last names from the current row
    if row[0] is not None and isinstance(row[0], str):
            # Extract first and last names from the current row
            # first_name, last_name = map(str.lower, row[0].split())
            names = row[0].split()
            first_name = names[0].lower() if names else ""
            last_name = names[1].lower() if len(names) > 1 else ""
            # Create the formatted email address with the first initial
            if flnames == '{first}.{last}':
                email = f"{first_name}.{last_name}{brands}"

            if flnames == '{f}.{last}':
                email = f"{first_name[0]}.{last_name}{brands}"

            if flnames == '{f}{last}':
                email = f"{first_name[0]}{last_name}{brands}"

            if flnames == '{first}_{last}':
                email = f"{first_name}_{last_name}{brands}"

            if flnames == '{first}':
                email = f"{first_name}{brands}"

            if flnames == '{first}{last}':
                email = f"{first_name}{last_name}{brands}"

            if flnames == '{first}{l}':
                email = f"{first_name}{last_name[0]}{brands}"

            if flnames == '{last}{f}':
                
                email = f"{last_name}{first_name[0]}{brands}"                                                
            # Append the formatted email to the list
            formatted_emails.append(email)

# Print or use the formatted email addresses as needed

# Create workbook and worksheet
wb = Workbook()
ws = wb.active

# Write header row
ws.cell(row=1, column=1).value = "Email Address"

# Write emails to cells
row_number = ws.max_row + 1 if ws.max_row > 0 else 1
for email in formatted_emails:
    ws.cell(row=row_number, column=1).value = email
    row_number += 1

# Save the workbook
wb.save("emails.xlsx")
# for email in formatted_emails:
#     print(email)
# import openpyxl

# # Load the Excel workbook
# workbook = openpyxl.load_workbook('data.xlsx')

# # Assuming the data is in the first sheet, change it accordingly if it's in a different sheet
# sheet = workbook.active

# # Create a new column header for the formatted emails
# sheet.cell(row=1, column=sheet.max_column + 1, value='Formatted Email')

# # Iterate through rows, starting from the second row (assuming the first row contains headers)
# for row in sheet.iter_rows(min_row=2, values_only=True):
#     # Extract first and last names from the current row
#     first_name, last_name = map(str.lower, row[0].split())

#     # Create the formatted email address with the first initial
#     email = f"{first_name[0]}.{last_name}@alexandermcqueen.com"

#     # Write the formatted email to the new column
#     sheet.cell(row=row[0], column=sheet.max_column, value=email)

# # Save the modified workbook to a new Excel file
# workbook.save('reports.xlsx')



